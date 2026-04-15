"""Migration report generator (Excel workbook).

Produces ``output/<workbook>/migration_report.xlsx`` at the end of every
pipeline run. An Excel workbook is the right format for this report
because reviewers want to sort, filter, and scan the Tableau ->
Power BI mapping — none of which Markdown does well.

Sheets:

- **Summary**           — Tableau vs PBI counts for every item type
                          (datasources, calcs, worksheets, dashboards, …).
- **Data Model**        — each Tableau datasource mapped to a PBI table,
                          with the table's kind (CSV import / empty
                          logical wrapper / calculated / PostgreSQL) and
                          source file + byte size.
- **Relationships**     — every relationship in the model, with
                          cardinality, direction, and active flag.
- **Visual Mapping**    — the main deliverable: one row per field-role
                          assignment, so you can filter by PBI visual
                          type to see "all my cards", "all my maps",
                          etc. Includes Tableau mark type, shelf
                          structure, and the PBI visualType + role +
                          table + column + aggregation.
- **Dashboards → Pages**— dashboard-to-page mapping with visual /
                          slicer / text / image counts.
- **Calculations → DAX**— every Tableau formula side-by-side with the
                          generated DAX and its kind (measure / calc
                          column).
- **Parameters**        — Tableau parameters with type, current value,
                          and allowable values.

All header rows are styled and frozen; columns are auto-sized to the
longest value (clamped) so the workbook is readable on open.
"""

import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")     # dark blue
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, color="1F4E78", size=14)
_SUBTLE_FONT = Font(italic=True, color="595959", size=10)
_ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")      # light grey
_WARN_FILL = PatternFill("solid", fgColor="FFF3CD")       # soft yellow
_EMPTY_FILL = PatternFill("solid", fgColor="F8D7DA")      # soft red
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")
_THIN = Side(border_style="thin", color="BFBFBF")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Reverse map of PBI aggregation-function codes used in prototypeQuery.
_AGG_NAME_BY_CODE = {
    0: "Sum", 1: "Avg", 2: "Count", 3: "Min", 4: "Max", 5: "CountD",
}


# ---------------------------------------------------------------------
# Mapping fidelity classifier
# ---------------------------------------------------------------------
#
# Three categories:
#   DIRECT        — Tableau construct has a 1:1 equivalent in PBI; migration
#                   should look and behave the same without tweaking.
#   SIMILAR       — PBI has a close approximation, but styling / encoding
#                   semantics differ enough that a human should eyeball it.
#   NOT_AVAILABLE — No native PBI equivalent; the visual needs to be
#                   rebuilt (custom visual, redesign, or replace with a
#                   different chart type).
#
# A reviewer only has to look at SIMILAR + NOT_AVAILABLE — those are the
# "manual review" buckets surfaced in the Summary and Manual Review sheets.

# PBI visuals that are a straight equivalent of a standard Tableau mark.
_DIRECT_VISUALS = {
    "clusteredBarChart", "clusteredColumnChart",
    "stackedBarChart", "stackedColumnChart",
    "hundredPercentStackedBarChart", "hundredPercentStackedColumnChart",
    "barChart", "columnChart",          # legacy aliases
    "lineChart",
    "areaChart", "stackedAreaChart", "hundredPercentStackedAreaChart",
    "pieChart", "donutChart",
    "scatterChart",
    "tableEx", "pivotTable",
    "matrix",
    "card", "multiRowCard", "kpi",
    "textbox",
    "slicer",
    "image",
}

# PBI visuals that approximate but do not perfectly reproduce Tableau
# behavior — always flag for review.
_SIMILAR_VISUALS = {
    "map", "filledMap", "shapeMap",
    "treemap",
    "lineClusteredColumnComboChart", "lineStackedColumnComboChart",
    "gauge",
    "funnel",
    "waterfall",
    "ribbonChart",
    "basicShape",
    "histogram",
}

# Tableau mark types with no native PBI equivalent. Presence of any of
# these in the worksheet context overrides the visualType classification.
_NOT_AVAILABLE_MARKS = {
    "GanttBar",          # Gantt bars — no native PBI Gantt
    "Polygon",           # Custom polygon marks (not geographic Multipolygon)
    "CircleView",        # Packed bubbles / circle views
    "PackedBubble",
    "Boxplot",
}

# Tableau marks that, even when mapped to a valid PBI visual, retain
# styling / encoding nuances the automated pipeline cannot fully carry.
# Forces a SIMILAR classification.
_REVIEW_MARKS = {
    "Shape",             # Custom shapes lost
    "Square",            # Heatmap / density — color ramp differs
    "Circle",            # Bubble / packed-bubble nuances
    "Multipolygon",      # Geographic fill style differs slightly
    "Automatic",         # Tableau auto-inference — verify PBI picked right
}


def _classify_mapping(marks, vtype):
    """Return ``(category, reason)`` for one worksheet's mapping.

    ``marks`` is the list of Tableau mark types on the worksheet;
    ``vtype`` is the PBI visualType chosen by the visual migrator.
    ``category`` ∈ ``{"DIRECT", "SIMILAR", "NOT_AVAILABLE"}``.
    ``reason`` is a short note suitable for the report cell.
    """
    marks = [m for m in (marks or []) if m]
    if not vtype or vtype == "(skipped)":
        return ("NOT_AVAILABLE",
                "No PBI visual was produced — the visual migrator could "
                "not map this worksheet.")

    # Any mark with no PBI equivalent wins.
    for m in marks:
        if m in _NOT_AVAILABLE_MARKS:
            return ("NOT_AVAILABLE",
                    f"Tableau '{m}' mark has no native Power BI equivalent "
                    f"— rebuild with a custom visual.")

    # Unrecognised PBI type → treat as SIMILAR to be safe.
    if vtype not in _DIRECT_VISUALS and vtype not in _SIMILAR_VISUALS:
        return ("SIMILAR",
                f"Uncommon PBI visualType '{vtype}' — verify against "
                f"Tableau intent.")

    if vtype in _SIMILAR_VISUALS:
        return ("SIMILAR",
                f"PBI '{vtype}' is the closest equivalent but usually "
                f"needs styling / encoding tweaks.")

    # Mark-driven review flag even when the PBI side is "direct".
    for m in marks:
        if m in _REVIEW_MARKS:
            return ("SIMILAR",
                    f"Tableau '{m}' mark mapped to '{vtype}' — review "
                    f"the encoding / styling fidelity.")

    return ("DIRECT", "Standard 1:1 mapping.")


# Row-shading colours per category.
_CAT_FILL = {
    "DIRECT":        PatternFill("solid", fgColor="E2F0D9"),   # soft green
    "SIMILAR":       PatternFill("solid", fgColor="FFF3CD"),   # soft yellow
    "NOT_AVAILABLE": PatternFill("solid", fgColor="F8D7DA"),   # soft red
}


# ---------------------------------------------------------------------
# Sheet writing helpers
# ---------------------------------------------------------------------

def _write_title(ws, title, subtitle=None):
    """Write a sheet title + optional subtitle in rows 1–2, return next free row."""
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = _SUBTLE_FONT
        return 4
    return 3


def _write_table(ws, start_row, headers, rows, row_fills=None):
    """Write a header row + data rows starting at ``start_row``.

    ``row_fills`` is an optional parallel list of PatternFills (or None)
    for per-row shading (used to flag empty-wrapper / warning rows).
    Applies zebra striping over that. Header row is frozen.
    """
    # Header
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = _CELL_BORDER

    # Data rows
    for ri, row in enumerate(rows):
        base_fill = row_fills[ri] if row_fills and ri < len(row_fills) else None
        stripe = _ZEBRA_FILL if (ri % 2 == 1 and base_fill is None) else None
        fill = base_fill or stripe
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=start_row + 1 + ri, column=ci, value=val)
            cell.alignment = _WRAP
            cell.border = _CELL_BORDER
            if fill:
                cell.fill = fill

    # Freeze below the header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate

    # Enable autofilter on the data range
    last_col = get_column_letter(len(headers))
    last_row = start_row + len(rows)
    ws.auto_filter.ref = f"A{start_row}:{last_col}{last_row}"


def _autosize_columns(ws, start_row, min_width=8, max_width=80):
    """Approximate auto-sizing based on the content in each column."""
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        longest = min_width
        for row in range(start_row, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            # Treat multi-line values conservatively: take the longest line.
            lines = str(val).splitlines() or [""]
            length = max(len(line) for line in lines)
            longest = max(longest, length + 2)
        ws.column_dimensions[letter].width = min(longest, max_width)


def _normalise(val):
    """Coerce values into something openpyxl can write safely."""
    if val is None:
        return ""
    if isinstance(val, (str, int, float, bool)):
        return val
    try:
        return json.dumps(val, ensure_ascii=False)
    except Exception:
        return str(val)


# ---------------------------------------------------------------------
# Section builders — each writes to its own sheet.
# ---------------------------------------------------------------------

def _classify_all_visuals(visual_map, ws_contexts):
    """Return ``[(worksheet_name, marks, vtype, category, reason), ...]``
    in a stable order, classifying every worksheet in ``visual_map`` +
    any that only appear in ``ws_contexts``.
    """
    ctx_by_name = {c.get("name", ""): c for c in (ws_contexts or [])}
    ordered = list((visual_map or {}).keys()) + [
        n for n in ctx_by_name if n not in (visual_map or {})
    ]
    seen = set()
    ordered = [n for n in ordered if not (n in seen or seen.add(n))]

    out = []
    for ws_name in ordered:
        ctx = ctx_by_name.get(ws_name, {})
        marks = ctx.get("mark_types", []) or []
        vis = (visual_map or {}).get(ws_name)
        vtype = (vis.get("visualType", "") if vis else "(skipped)") or "(skipped)"
        category, reason = _classify_mapping(marks, vtype)
        out.append((ws_name, marks, vtype, category, reason))
    return out


def _sheet_summary(wb, workbook_name, input_path, metadata, bim, visual_map,
                   pages, db_contexts, stats, now):
    ws = wb.create_sheet("Summary")
    row = _write_title(
        ws,
        f"Migration report — {workbook_name}",
        f"Source: {input_path or '(unknown)'}    Generated: {now}    Target: Power BI (PBIP / model.bim)",
    )

    model = bim.get("model", {}) if bim else {}
    tables = [t for t in model.get("tables", []) if not t.get("isPrivate")]
    total_measures = sum(len(t.get("measures", []) or []) for t in tables)
    total_calc_cols = sum(
        sum(1 for c in (t.get("columns", []) or []) if c.get("type") == "calculated")
        for t in tables
    )
    total_cols = sum(len(t.get("columns", []) or []) for t in tables) - total_calc_cols
    rels = model.get("relationships", []) or []

    headers = ["Item", "Tableau (source)", "Power BI (target)"]
    rows = [
        ["Workbook", os.path.basename(input_path or ""), workbook_name or ""],
        ["Datasources / Tables",
         len(metadata.get("datasources", []) or []), len(tables)],
        ["Physical columns", "—", total_cols],
        ["Calculations",
         len(metadata.get("calculations", []) or []),
         f"{total_measures} measures + {total_calc_cols} calc columns"],
        ["Parameters",
         len(metadata.get("parameters", []) or []),
         "Parameters table (measures)"],
        ["Worksheets / Visuals",
         len(metadata.get("worksheets", []) or []), len(visual_map or {})],
        ["Dashboards / Pages",
         len(db_contexts or []), len(pages or [])],
        ["Relationships", "—", len(rels)],
    ]
    _write_table(ws, row, headers, [[_normalise(c) for c in r] for r in rows])

    # Pipeline stats block below
    if stats:
        row2 = ws.max_row + 3
        ws.cell(row=row2, column=1, value="Pipeline stats").font = _TITLE_FONT
        _write_table(
            ws, row2 + 1, ["Metric", "Value"],
            [[_normalise(k), _normalise(v)] for k, v in stats.items()],
        )

    _autosize_columns(ws, start_row=3)


def _pct(n, total):
    if not total:
        return "0%"
    return f"{(n * 100 / total):.0f}%"


def _sheet_data_model(wb, metadata, bim, csv_dir):
    ws = wb.create_sheet("Data Model")
    row = _write_title(
        ws, "Data model — Tableau datasources → Power BI tables",
        "Kind: CSV import / Empty wrapper (Tableau blend) / PostgreSQL / Calculated",
    )

    model = bim.get("model", {}) if bim else {}
    tables = [t for t in model.get("tables", []) if not t.get("isPrivate")]
    ds_by_caption = {ds.get("caption", ""): ds for ds in (metadata.get("datasources", []) or [])}

    headers = ["Tableau Datasource", "PBI Table", "Columns", "Kind", "Source file", "Size (bytes)", "Notes"]
    data_rows = []
    row_fills = []

    for t in tables:
        name = t.get("name", "")
        cols = t.get("columns", []) or []
        n_cols = len(cols)
        parts = t.get("partitions", []) or []
        part_expr = ""
        if parts:
            exp = parts[0].get("source", {}).get("expression", "")
            part_expr = "\n".join(exp) if isinstance(exp, list) else str(exp)

        size_bytes = ""
        source_file = ""
        notes = ""
        fill = None

        if "#table(" in part_expr and "{}" in part_expr:
            kind = "Empty (logical wrapper)"
            notes = "No physical extract — Tableau blend / Multiple Connections. Calculated columns and measures still evaluate against related tables."
            fill = _WARN_FILL
        elif "Csv.Document" in part_expr:
            kind = "CSV import"
            source_file = f"{name}.csv"
            csv_path = os.path.join(csv_dir, source_file) if csv_dir else ""
            if csv_path and os.path.exists(csv_path):
                size_bytes = os.path.getsize(csv_path)
            else:
                fill = _EMPTY_FILL
                notes = "CSV expected but missing on disk"
        elif "PostgreSQL.Database" in part_expr:
            kind = "PostgreSQL"
            notes = "Loads from external PG database (see M expression)"
        elif part_expr.strip().startswith("ROW("):
            kind = "Calculated (Parameters)"
            notes = "DAX ROW(...) — Tableau parameters"
        else:
            kind = "Calculated / other"
            notes = (part_expr[:100] + "…") if len(part_expr) > 100 else part_expr

        tab_src = name if name in ds_by_caption else "(derived)"
        data_rows.append([tab_src, name, n_cols, kind, source_file, size_bytes, notes])
        row_fills.append(fill)

    _write_table(ws, row, headers, data_rows, row_fills)
    _autosize_columns(ws, start_row=row)


def _sheet_relationships(wb, bim):
    ws = wb.create_sheet("Relationships")
    row = _write_title(ws, "Relationships", "Model relationships emitted into model.bim")

    rels = (bim.get("model", {}) if bim else {}).get("relationships", []) or []
    headers = ["From Table", "From Column", "To Table", "To Column",
               "From Card.", "To Card.", "Direction", "Active"]
    data_rows = []
    row_fills = []

    for r in rels:
        active = r.get("isActive", True)
        data_rows.append([
            r.get("fromTable", ""),
            r.get("fromColumn", ""),
            r.get("toTable", ""),
            r.get("toColumn", ""),
            r.get("fromCardinality", "many"),
            r.get("toCardinality", "one"),
            r.get("crossFilteringBehavior", "oneDirection") or "oneDirection",
            "yes" if active else "no",
        ])
        row_fills.append(None if active else _WARN_FILL)

    _write_table(ws, row, headers, data_rows, row_fills)
    _autosize_columns(ws, start_row=row)


def _sheet_visual_mapping(wb, visual_map, ws_contexts, classification):
    """One row per Tableau worksheet's field-role assignment, plus two
    columns classifying the mapping fidelity so reviewers can filter.

    Filtering by ``Mapping Fidelity = SIMILAR`` or ``NOT_AVAILABLE`` gives
    you the manual-review worklist. ``classification`` is the list
    produced by ``_classify_all_visuals``.
    """
    ws = wb.create_sheet("Visual Mapping")
    row = _write_title(
        ws, "Visual mapping — Tableau worksheets → Power BI visuals",
        "One row per field-role. Filter 'Mapping Fidelity' to see "
        "DIRECT / SIMILAR (manual review) / NOT_AVAILABLE (rebuild).",
    )

    ctx_by_name = {c.get("name", ""): c for c in (ws_contexts or [])}
    cat_by_name = {c[0]: (c[3], c[4]) for c in (classification or [])}

    headers = ["#", "Tableau Worksheet", "Tableau Mark", "Shelf Structure",
               "PBI Visual Type", "Mapping Fidelity", "Review Note",
               "Role", "Table", "Column", "Is Measure?", "Aggregation"]
    data_rows = []
    row_fills = []

    # Use the same ordering as classification so the two views line up.
    ordered = [c[0] for c in (classification or [])]
    if not ordered:
        ordered = list((visual_map or {}).keys()) + [
            n for n in ctx_by_name if n not in (visual_map or {})
        ]
        seen = set()
        ordered = [n for n in ordered if not (n in seen or seen.add(n))]

    for i, ws_name in enumerate(ordered, start=1):
        ctx = ctx_by_name.get(ws_name, {})
        mark = ", ".join(ctx.get("mark_types", []) or []) or "—"
        shelf = ctx.get("shelf_structure") or ""
        if not isinstance(shelf, str):
            n_rows = len(ctx.get("rows_fields") or [])
            n_cols = len(ctx.get("cols_fields") or [])
            shelf = f"rows({n_rows}) × cols({n_cols})"

        category, reason = cat_by_name.get(ws_name, ("", ""))
        cat_fill = _CAT_FILL.get(category)

        vis = (visual_map or {}).get(ws_name)
        if not vis:
            data_rows.append([i, ws_name, mark, shelf,
                              "(skipped)", category, reason,
                              "—", "—", "—", "—", "—"])
            row_fills.append(cat_fill or _EMPTY_FILL)
            continue

        vtype = vis.get("visualType", "")
        assignments = _list_field_assignments(vis)
        if not assignments:
            data_rows.append([i, ws_name, mark, shelf, vtype,
                              category, reason,
                              "(no fields)", "—", "—", "—", "—"])
            row_fills.append(cat_fill or _WARN_FILL)
            continue

        for j, a in enumerate(assignments):
            data_rows.append([
                i if j == 0 else "",
                ws_name if j == 0 else "",
                mark if j == 0 else "",
                shelf if j == 0 else "",
                vtype if j == 0 else "",
                category if j == 0 else "",
                reason if j == 0 else "",
                a["role"], a["table"], a["column"],
                "yes" if a["is_measure"] else "no",
                a["aggregation"],
            ])
            # Shade the first row of each worksheet group by category;
            # subsequent field rows stay unshaded so the grouping is
            # visually obvious.
            row_fills.append(cat_fill if j == 0 else None)

    _write_table(ws, row, headers, data_rows, row_fills)
    _autosize_columns(ws, start_row=row)


def _list_field_assignments(single_visual):
    """Extract (role, table, column, is_measure, aggregation) tuples."""
    out = []

    # Preferred: projections + prototypeQuery.Select give us roles + bindings.
    projections = single_visual.get("projections", {}) or {}
    pq = single_visual.get("prototypeQuery", {}) or {}
    selects = pq.get("Select", []) or []

    # Map queryRef -> select spec to resolve role -> field binding
    select_by_ref = {}
    for s in selects:
        if not isinstance(s, dict):
            continue
        name = s.get("Name", "")
        if name:
            select_by_ref[name] = s

    for role, items in projections.items():
        for it in (items or []):
            qref = it.get("queryRef", "")
            spec = select_by_ref.get(qref, {})
            entry = _describe_select(spec, fallback_ref=qref)
            entry["role"] = role
            out.append(entry)

    # Fallback: no projections — walk selects and use role hints on the select.
    if not out and selects:
        for s in selects:
            entry = _describe_select(s)
            if entry["column"]:
                out.append(entry)

    return out


def _describe_select(s, fallback_ref=""):
    """Return {role, table, column, is_measure, aggregation} best-effort."""
    entry = {"role": "", "table": "", "column": "", "is_measure": False,
             "aggregation": ""}
    if not isinstance(s, dict):
        if fallback_ref:
            entry["column"] = fallback_ref
        return entry
    try:
        if "Aggregation" in s:
            agg = s["Aggregation"]
            inner = agg.get("Expression", {}).get("Column", {})
            entry["table"] = inner.get("Expression", {}).get(
                "SourceRef", {}).get("Entity", "")
            entry["column"] = inner.get("Property", "")
            code = agg.get("Function", -1)
            entry["aggregation"] = _AGG_NAME_BY_CODE.get(code, f"Agg{code}" if code != -1 else "")
            entry["is_measure"] = False
            return entry
        if "Column" in s:
            c = s["Column"]
            entry["table"] = c.get("Expression", {}).get(
                "SourceRef", {}).get("Entity", "")
            entry["column"] = c.get("Property", "")
            entry["is_measure"] = False
            return entry
        if "Measure" in s:
            m = s["Measure"]
            entry["table"] = m.get("Expression", {}).get(
                "SourceRef", {}).get("Entity", "")
            entry["column"] = m.get("Property", "")
            entry["is_measure"] = True
            return entry
    except Exception:
        pass
    if fallback_ref:
        entry["column"] = fallback_ref
    return entry


def _sheet_dashboards(wb, pages, db_contexts):
    ws = wb.create_sheet("Dashboards")
    row = _write_title(
        ws, "Dashboards → Report pages",
        "Per-page counts of chart visuals, slicers, text boxes, and images.",
    )

    headers = ["Tableau Dashboard", "PBI Page", "Chart Visuals",
               "Slicers", "Text Boxes", "Images", "Notes"]
    data_rows = []
    row_fills = []
    covered = set()

    for p in pages or []:
        disp = p.get("displayName") or p.get("name", "")
        covered.add(disp)
        vcs = p.get("visualContainers", []) or []
        n_vis = n_slicer = n_text = n_img = 0
        for vc in vcs:
            cfg_raw = vc.get("config", "")
            try:
                cfg = json.loads(cfg_raw) if isinstance(cfg_raw, str) else cfg_raw
            except Exception:
                cfg = {}
            sv = cfg.get("singleVisual", {}) if isinstance(cfg, dict) else {}
            vt = sv.get("visualType", "")
            if vt == "slicer":
                n_slicer += 1
            elif vt == "textbox":
                n_text += 1
            elif vt == "image":
                n_img += 1
            elif vt:
                n_vis += 1
        data_rows.append([disp, disp, n_vis, n_slicer, n_text, n_img, ""])
        row_fills.append(None)

    for db in (db_contexts or []):
        name = db.get("name", "")
        if name in covered:
            continue
        data_rows.append([name, "(not migrated)",
                          len(db.get("zones", []) or []), 0, 0, 0,
                          "Dashboard has no corresponding PBI page"])
        row_fills.append(_WARN_FILL)

    _write_table(ws, row, headers, data_rows, row_fills)
    _autosize_columns(ws, start_row=row)


def _sheet_calculations(wb, metadata, bim):
    ws = wb.create_sheet("Calculations")
    row = _write_title(
        ws, "Calculations — Tableau formulas → DAX",
        "Each Tableau calc side-by-side with its generated DAX.",
    )

    calcs = metadata.get("calculations", []) or []

    dax_by_caption = {}
    for t in (bim.get("model", {}) if bim else {}).get("tables", []) or []:
        for m in t.get("measures", []) or []:
            n = m.get("name")
            if n:
                dax_by_caption.setdefault(n, ("measure", m.get("expression", ""), t.get("name", "")))
        for c in t.get("columns", []) or []:
            if c.get("type") == "calculated":
                n = c.get("name")
                if n:
                    dax_by_caption.setdefault(n, ("calc column", c.get("expression", ""), t.get("name", "")))

    headers = ["#", "Tableau Field", "Kind", "PBI Table",
               "Tableau Formula", "DAX Expression"]
    data_rows = []
    row_fills = []
    for i, c in enumerate(calcs, start=1):
        caption = c.get("caption", c.get("name", ""))
        kind, dax, tbl = dax_by_caption.get(caption, ("(not emitted)", "", ""))
        data_rows.append([
            i, caption, kind, tbl,
            (c.get("formula", "") or "").strip(),
            (dax or "").strip(),
        ])
        row_fills.append(_WARN_FILL if kind == "(not emitted)" else None)

    _write_table(ws, row, headers, data_rows, row_fills)
    # Formula + DAX columns wrap to ~90 chars
    for letter in ("E", "F"):
        ws.column_dimensions[letter].width = 90
    _autosize_columns(ws, start_row=row, min_width=8, max_width=90)


def _sheet_parameters(wb, metadata):
    params = metadata.get("parameters", []) or []
    if not params:
        return
    ws = wb.create_sheet("Parameters")
    row = _write_title(ws, "Parameters",
                       "Tableau parameters carried over as measures in the Parameters table.")

    headers = ["Name", "Data type", "Current value", "Allowable values"]
    data_rows = []
    for p in params:
        allow = p.get("allowable_values") or []
        if isinstance(allow, list):
            if len(allow) > 10:
                allow_str = ", ".join(str(a) for a in allow[:10]) + f"  (+{len(allow) - 10} more)"
            else:
                allow_str = ", ".join(str(a) for a in allow)
        else:
            allow_str = str(allow)
        data_rows.append([
            p.get("name", ""), p.get("datatype", ""),
            str(p.get("current_value", "")), allow_str,
        ])

    _write_table(ws, row, headers, data_rows)
    _autosize_columns(ws, start_row=row)


# ---------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------

def generate_migration_report(
    output_dir,
    workbook_name,
    input_path,
    metadata,
    bim,
    visual_map=None,
    pages=None,
    ws_contexts=None,
    db_contexts=None,
    csv_dir=None,
    warnings=None,
    stats=None,
):
    """Write ``output_dir/migration_report.xlsx`` and return its path.

    Missing optional data (e.g. ``visual_map=None``) results in the
    corresponding sheet being omitted rather than raising.
    """
    os.makedirs(output_dir, exist_ok=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()
    # Remove the default blank sheet — we'll create named sheets.
    wb.remove(wb.active)

    # Compute fidelity classification once — the Visual Mapping sheet consumes it.
    classification = _classify_all_visuals(visual_map, ws_contexts) \
        if (visual_map or ws_contexts) else []

    _sheet_summary(wb, workbook_name, input_path, metadata, bim, visual_map,
                   pages, db_contexts, stats, now)
    _sheet_data_model(wb, metadata, bim, csv_dir)
    _sheet_relationships(wb, bim)
    if visual_map or ws_contexts:
        _sheet_visual_mapping(wb, visual_map, ws_contexts, classification)
    if pages or db_contexts:
        _sheet_dashboards(wb, pages, db_contexts)
    _sheet_calculations(wb, metadata, bim)
    _sheet_parameters(wb, metadata)

    if warnings:
        ws = wb.create_sheet("Warnings")
        _write_title(ws, "Warnings & notes",
                     "Pipeline observations worth surfacing to a reviewer.")
        _write_table(ws, 3, ["#", "Message"],
                     [[i + 1, _normalise(w)] for i, w in enumerate(warnings)])
        _autosize_columns(ws, start_row=3)

    path = os.path.join(output_dir, "migration_report.xlsx")
    wb.save(path)
    return path
